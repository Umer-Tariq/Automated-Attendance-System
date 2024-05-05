import openpyxl
import datetime
from openpyxl.utils import get_column_letter
import pandas as pd
import random

def get_column_values(sheet, column_number):
    column_values = []
    for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=column_number, max_col=column_number):
        cell_value = row[0].value
        column_values.append(cell_value)
    return column_values

def get_excel_cell(row, col):
    col_letters = ""
    while col >= 0:
        col_letters = chr(65 + col % 26) + col_letters
        col = col // 26 - 1 if col >= 26 else -1
    row_number = row + 1
    cell_ref = f"{col_letters}{row_number}"
    return cell_ref

def read_file(room):
    wb = openpyxl.load_workbook('timetable.xlsx')
    current_day_name = datetime.datetime.now().strftime('%A')
    #current_day_name = 'WEDNESdAY'
    sheet_num = None
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        first_row = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
        for cell_value in first_row:
            weekday_name = cell_value
            for s in weekday_name:
                s = str(s)
                if current_day_name.lower() in s.lower():
                    print('Found matching day:', current_day_name)
                    sheet_num = sheet  
                    break
            if sheet_num:  
                break
        if sheet_num:  
            break
    second_row_values = [cell.value for cell in sheet_num[2]]
    current_time_12hr = datetime.datetime.now().strftime('%I:%M %p')
    current_hour = current_time_12hr.split(':')[0] 
    if(current_hour[0] == '0'):
        curr = ''
        for i in range(0, len(current_hour)):
            if current_hour[i] != '0':
                break
            else:
                curr = current_hour[i + 1:len(current_hour)]
        current_hour = curr
    index = 1
    for i in second_row_values:
        if current_hour == str(i).split(':')[0]:
            break
        index += 1
    column_values = get_column_values(sheet, index)

    item = str
    for i in column_values:

        if room in i:
           item = i 
    parts = item.split(' ')
    return parts[1].split('-')[1], parts[2], parts[3]

def find_cell_num( file_name, var ):
    cell_coordinates = None
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == var:
                cell_coordinates = cell.coordinate
                break
        if cell_coordinates:
            break
    
    return cell_coordinates

def extract_row_or_col(col_reference, option):
    ccol = ''
    crow = ''
    for i in range(0, len(col_reference)):
        if col_reference[i] >= '0' and col_reference[i] <= '9':
            crow += col_reference[i]
        else:
            ccol += col_reference[i]
    if option :
        return ccol
    else:
        return crow

def mark_attendance(file_name):
    df = pd.read_excel('6K_CS-4002.xlsx')
    print(df)
    roll_numbers_list = df['Roll_Number'].tolist()
    print(roll_numbers_list)
    incoming_student = []
    current_time_12hr = datetime.datetime.now().strftime('%I:%M %p')
    current_hour = current_time_12hr.split(':')[0] 
    if(current_hour[0] == '0'):
        curr = ''
        for i in range(0, len(current_hour)):
            if current_hour[i] != '0':
                break
            else:
                curr = current_hour[i + 1:len(current_hour)]
        current_hour = curr
    var = datetime.datetime.now().strftime('%Y-%m-%d') + ' ' + 'time'
    print(var)
    coordinates = find_cell_num('6K_CS-4002.xlsx', var)
    col_reference = extract_row_or_col(coordinates, 1)
    var = datetime.datetime.now().strftime('%Y-%m-%d') + ' ' + 'status'
    coordinates = find_cell_num('6K_CS-4002.xlsx', var)
    col2_reference = extract_row_or_col(coordinates, 1)
    len_of_children = len(roll_numbers_list)

    while len(incoming_student) != len_of_children:
        tries = 0
        value_to_find = random.choice(roll_numbers_list)
        roll_numbers_list.remove(value_to_find)
        print(value_to_find)
        incoming_student.append(value_to_find)
        cell_location = find_cell_num('6K_CS-4002.xlsx',value_to_find )
        while tries < 3:
            scanned_status = random.randint(0,1)
            if scanned_status:
                row_number = extract_row_or_col(cell_location, 0)
                print("Cell reference:", cell_location)
                cell_num1 = col_reference + row_number
                cell_num2 = col2_reference + row_number
                print('Attendance Marked! ' + str(value_to_find)  + '   pos =  ' +  str(cell_num1) )
                ''''wb = openpyxl.load_workbook('6K_CS-4002.xlsx')
                sheet = wb.active
                record_time = datetime.datetime.now().strftime('%I:%M %p')
                record_mins = record_time.split(':')[1].split(' ')[0]
                if( record_mins <= 15 ):
                    status = 'Present'
                elif( record_mins > 15 and record_mins <= 29 ):
                    status = 'Late'
                else:
                    status = 'Absent'
                wb[cell_num1] = record_time
                wb[cell_num1] = status
                wb.save('6K_CS-4002.xlsx')'''
                break
            else:
                print('Couldnt Detect. Pls tri again!')
                tries += 1

        if  tries >= 3 :
            print('Please use biomeric sensing') 
    

def prepare_file(sec, ccode):
    file_name = sec + '_' + ccode + '.xlsx'
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.active
    print(sheet.max_column)
    last_column_number = sheet.max_column + 1
    last_column_letter = get_column_letter(last_column_number)
    time_cell_number = last_column_letter + '1'
    status_cell_number = get_column_letter(last_column_number + 1) + '1'
    today_date_time = datetime.datetime.now().strftime('%Y-%m-%d') + ' ' + 'time'
    sheet[time_cell_number] = today_date_time
    print(time_cell_number)
    last_column_number = sheet.max_column + 1
    last_column_letter = get_column_letter(last_column_number)
    time_cell_number = last_column_letter + '1'
    today_date_time = datetime.datetime.now().strftime('%Y-%m-%d') + ' ' + 'status'
    sheet[time_cell_number] = today_date_time
    wb.save(file_name)
    mark_attendance(file_name)

#section, subject, ccode =  read_file('R11')
#prepare_file(section, ccode)
mark_attendance('ccc')


